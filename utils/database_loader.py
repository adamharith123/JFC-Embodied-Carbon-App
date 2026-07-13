from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import os


from utils.constants import (
    CARBON_DATABASE_FILE,
    STANDARDS_DATABASE_FILE,
    USER_INPUT_DATABASE_FILE,
)

# ==========================================================
# Generic Workbook Functions
# ==========================================================

def workbook_exists(path: Path) -> bool:
    """Return True if the workbook exists."""

    return path.exists() and path.suffix.lower() in [
        ".xlsx",
        ".xlsm",
        ".xls",
    ]

def _file_mtime(path):
    """
    Returns the file's last-modified time, used purely as a cache key
    so cached data automatically refreshes if the Excel file changes
    on disk, without needing a manual cache-clear.
    """
    try:
        return os.path.getmtime(path)
    except OSError:
        return None


def get_workbook_sheet_names(path: Path):
    """Return worksheet names."""

    if not workbook_exists(path):
        return []

    wb = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    return wb.sheetnames


def load_sheet(path: Path, sheet_name: str):
    """
    Load a worksheet into a pandas DataFrame.

    The BOQ worksheet has title rows above the
    column headings, so row 6 is used as the header.

    All other worksheets use the default header row.
    """

    if not workbook_exists(path):
        return pd.DataFrame()

    try:

        if sheet_name == "BOQ":

            return pd.read_excel(
                path,
                sheet_name=sheet_name,
                header=5,
            )
        if sheet_name == "Apparatus Output":

            return pd.read_excel(
                path,
                sheet_name=sheet_name,
                header=1,
            )

        return pd.read_excel(
            path,
            sheet_name=sheet_name,
        )

    except Exception:

        return pd.DataFrame()


# ==========================================================
# Carbon Database
# ==========================================================

@st.cache_data
def load_carbon_database(_mtime=None):
    """
    Load the embodied carbon database.

    Returns
    -------
    dict
        Dictionary containing all engineering datasets
        required by the application.
    """

    boq = load_sheet(
        CARBON_DATABASE_FILE,
        "BOQ",
    )

    apparatus_output = load_sheet(
        CARBON_DATABASE_FILE,
        "Apparatus Output",
    )

    dropdowns = load_sheet(
        CARBON_DATABASE_FILE,
        "dropdowns",
    )
    product_output = load_product_output()  

    systems = []

    # Prefer Apparatus Output because it is already
    # aggregated at apparatus level.

    if (
        not apparatus_output.empty
        and "Apparatus" in apparatus_output.columns
    ):

        systems = (
            apparatus_output["Apparatus"]
            .dropna()
            .astype(str)
            .sort_values()
            .unique()
            .tolist()
        )

    # Fallback to BOQ if required

    elif (
        not boq.empty
        and "Apparatus" in boq.columns
    ):

        systems = (
            boq["Apparatus"]
            .dropna()
            .astype(str)
            .sort_values()
            .unique()
            .tolist()
        )

    return {

        "path": CARBON_DATABASE_FILE,

        "exists": workbook_exists(
            CARBON_DATABASE_FILE
        ),

        "sheets": get_workbook_sheet_names(
            CARBON_DATABASE_FILE
        ),

        "boq": boq,

        "apparatus_output": apparatus_output,

        "dropdowns": dropdowns,

        "systems": systems,

        "product_output": product_output,

    }


# ==========================================================
# Standards Database
# ==========================================================

@st.cache_data
def load_standards_database(_mtime=None):
    """
    Load the standards database.
    """

    return {

        "path": STANDARDS_DATABASE_FILE,

        "exists": workbook_exists(
            STANDARDS_DATABASE_FILE
        ),

        "sheets": get_workbook_sheet_names(
            STANDARDS_DATABASE_FILE
        ),

        "standards_list": load_sheet(
            STANDARDS_DATABASE_FILE,
            "Standards List",
        ),

        "building_class": load_sheet(
            STANDARDS_DATABASE_FILE,
            "Building Class",
        ),

    }

# ==========================================================
# User Input Database
# ==========================================================

def load_user_input_database():
    """
    Load the engineering user input database.
    """

    return {

        "path": USER_INPUT_DATABASE_FILE,

        "exists": workbook_exists(
            USER_INPUT_DATABASE_FILE
        ),

        "sheets": get_workbook_sheet_names(
            USER_INPUT_DATABASE_FILE
        ),

        "database": load_sheet(
            USER_INPUT_DATABASE_FILE,
            "Database",
        ),

        "user_inputs": load_sheet(
            USER_INPUT_DATABASE_FILE,
            "User Inputs",
        ),

    }


# ==========================================================
# Application Loader
# ==========================================================

def load_all_databases():

    return {

        "carbon": load_carbon_database(_mtime=_file_mtime(STANDARDS_DATABASE_FILE)),

        "standards": load_standards_database(_mtime=_file_mtime(STANDARDS_DATABASE_FILE)),

        "user_input": load_user_input_database(),

    }


# ==========================================================
# Database Status
# ==========================================================

def database_status_summary():

    db = load_all_databases()

    return {

        "Carbon Database": {

            "exists": db["carbon"]["exists"],

            "path": db["carbon"]["path"],

            "sheets": db["carbon"]["sheets"],

            "rows": len(
                db["carbon"]["apparatus_output"]
            ),

        },

        "Standards Database": {

            "exists": db["standards"]["exists"],

            "path": db["standards"]["path"],

            "sheets": db["standards"]["sheets"],

            "rows": len(
                db["standards"]["standards_list"]
            ),

        },

    }
# ==========================================================
# Proposed Design Helpers
# ==========================================================

def get_building_classes():
    """
    Return the available NCC building classes.
    """

    df = load_standards_database(_mtime=_file_mtime(STANDARDS_DATABASE_FILE))["building_class"]

    if df.empty:
        return []

    first_col = df.columns[0]

    return (
        df[first_col]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

def load_product_output():
    """
    Load the Product Output sheet - branded/component-level carbon
    factors, filterable by Apparatus.
    """

    return load_sheet(
        CARBON_DATABASE_FILE,
        "Product Output",
    )


def get_standards_list():
    """
    Return the Standards List worksheet.
    """

    return load_standards_database(_mtime=_file_mtime(STANDARDS_DATABASE_FILE))["standards_list"]