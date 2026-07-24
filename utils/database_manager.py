"""
Database Manager - backing logic for the Help page's "Manage Database"
section.

Two-folder model:

- database/databases/  -> the LIVE file the app actually reads from
  (paths already defined in utils.constants - unchanged by this module).
- database/defaults/    -> a permanent copy of each file, seeded once
  from whatever is in database/databases/ the first time the app runs
  with this feature. The app never writes to this folder again after
  that - it is only ever read from, to support "revert to default".

Three actions, each a file copy plus (for uploads) a structural
validation pass:
  - download current   -> read bytes straight from the live path
  - upload replacement  -> validate, archive the old live file, then
                           overwrite the live path with the upload
  - revert to default   -> copy database/defaults/<file> back over the
                           live path, discarding whatever was uploaded

Validation is "structural minimum": required sheets must exist, and
the specific columns the app's code actually reads by name must be
present (verified against utils/database_loader.py,
utils/standards_engine.py, utils/ui_structure_loader.py and
utils/proposed_design_calculations.py). Extra sheets/columns are
allowed and ignored. This guarantees the app won't crash reading the
file - it does not guarantee the new data is engineering-correct.
"""

import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from utils.constants import (
    CARBON_DATABASE_FILE,
    STANDARDS_DATABASE_FILE,
    CALC_RULES_DATABASE_FILE,
    ARCHIVE_DIR,
    DEFAULT_DATABASE_DIR,
)


def _norm(name):
    """
    Normalise a column/sheet name for comparison - strips whitespace
    and non-breaking spaces (the Standards List / Building Class
    sheets use \xa0 in some of their headers).
    """
    return str(name).replace("\xa0", "").strip().casefold()


# ==========================================================
# Registry - one entry per managed workbook
# ==========================================================
# Each sheet maps to (header_row_index, [required column names]).
# header_row_index matches what utils/database_loader.py already
# uses when reading that sheet (e.g. "Apparatus Output" has a title
# row above the headers, so header=1).

DATABASE_REGISTRY = {
    "carbon": {
        "label": "Carbon Database (ARUP_v2_Finalised.xlsx)",
        "live_path": CARBON_DATABASE_FILE,
        "default_path": DEFAULT_DATABASE_DIR / CARBON_DATABASE_FILE.name,
        "sheets": {
            "Apparatus Output": (
                1,
                [
                    "Category", "System", "Apparatus", "Units",
                    "A1-3 (kg CO2e)", "A4 (kg CO2e)", "A5 (kg CO2e)",
                    ("Total (A1-3 + A4 + A5)", "Total GWP (A1-3 + A4 + A5)"),
                ],
            ),
        },
    },
    "standards": {
        "label": "Building Class Database (standards_database_master.xlsx)",
        "live_path": STANDARDS_DATABASE_FILE,
        "default_path": DEFAULT_DATABASE_DIR / STANDARDS_DATABASE_FILE.name,
        "sheets": {
            # Neither sheet's columns are read by name anywhere in the
            # app (both are accessed generically, e.g. "first column"),
            # so only sheet presence is required here.
            "Standards List": (0, []),
            "Building Class": (0, []),
        },
    },
    "calc": {
        "label": "UI & Calculation Database (Standards_Calc_Database_Finalised.xlsx)",
        "live_path": CALC_RULES_DATABASE_FILE,
        "default_path": DEFAULT_DATABASE_DIR / CALC_RULES_DATABASE_FILE.name,
        "sheets": {
            "ui_structure": (
                0,
                [
                    "Category", "Category Name", "Group", "Label", "Apparatus",
                    "Archetype", "Modes", "Allow Multiple Rows", "Units",
                    "Parent", "Linked Mode", "Formula System",
                    "Formula Component", "Formula Parameters",
                    "Counted Apparatus", "Requires FRL",
                ],
            ),
            "calc_rules": (
                0,
                ["system", "component", "parameter", "condition_value", "value"],
            ),
            "frl_reference": (
                0,
                [
                    "Apparatus", "Product Type", "FRL (min)",
                    "Thickness (mm)", "Layers", "Density (kg/m3)",
                ],
            ),
        },
    },
}


# ==========================================================
# One-time seeding of the permanent default copies
# ==========================================================

def ensure_defaults_seeded():
    """
    For each managed workbook, if a default copy doesn't exist yet,
    copy the current live file to database/defaults/. Never
    overwrites a default that's already there - safe to call on
    every page load.
    """
    DEFAULT_DATABASE_DIR.mkdir(parents=True, exist_ok=True)

    for entry in DATABASE_REGISTRY.values():
        default_path = entry["default_path"]
        live_path = entry["live_path"]

        if not default_path.exists() and live_path.exists():
            shutil.copy2(live_path, default_path)


# ==========================================================
# Validation
# ==========================================================

def validate_workbook(file_obj, db_key):
    """
    Structural-minimum validation against DATABASE_REGISTRY[db_key]:
    required sheets must exist, and each sheet's required columns
    must be present (extra sheets/columns are fine).

    file_obj may be a path or a file-like object (e.g. a Streamlit
    UploadedFile) - it is read multiple times, so it is rewound
    between reads.

    Returns (ok: bool, message: str).
    """
    entry = DATABASE_REGISTRY[db_key]

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as e:
        return False, f"Could not open file as an Excel workbook: {e}"

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    available_sheets = wb.sheetnames
    missing_sheets = [s for s in entry["sheets"] if s not in available_sheets]

    if missing_sheets:
        return False, f"Missing required sheet(s): {', '.join(missing_sheets)}"

    for sheet_name, (header_row, required_columns) in entry["sheets"].items():
        if not required_columns:
            continue

        try:
            df = pd.read_excel(file_obj, sheet_name=sheet_name, header=header_row)
        except Exception as e:
            return False, f"Could not read sheet '{sheet_name}': {e}"

        if hasattr(file_obj, "seek"):
            file_obj.seek(0)

        available_cols = {_norm(c) for c in df.columns}
        missing_cols = []
        for required in required_columns:
            # A required entry can be a single column name, or a
            # tuple/list of acceptable alternative names (any one
            # present satisfies the requirement) - mirrors the
            # tolerance utils/proposed_design_calculations.py's
            # _find_column() already has for renamed columns.
            alternatives = required if isinstance(required, (tuple, list)) else (required,)
            if not any(_norm(alt) in available_cols for alt in alternatives):
                missing_cols.append(" / ".join(alternatives))

        if missing_cols:
            return False, (
                f"Sheet '{sheet_name}' is missing required column(s): "
                f"{', '.join(missing_cols)}"
            )

    return True, "Validation passed."


# ==========================================================
# Actions
# ==========================================================

def get_live_bytes(db_key):
    """Return the current live file's bytes, or None if it doesn't exist."""
    live_path = DATABASE_REGISTRY[db_key]["live_path"]
    if not live_path.exists():
        return None
    return live_path.read_bytes()


def replace_live_database(db_key, uploaded_file):
    """
    Validates an uploaded file against db_key's schema. If it passes,
    archives the current live file (timestamped, in ARCHIVE_DIR) then
    overwrites the live file with the upload.

    Returns (ok: bool, message: str).
    """
    ok, message = validate_workbook(uploaded_file, db_key)

    if not ok:
        return False, message

    entry = DATABASE_REGISTRY[db_key]
    live_path = entry["live_path"]

    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

    if live_path.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archived_name = f"{live_path.stem}_{timestamp}{live_path.suffix}"
        shutil.copy2(live_path, ARCHIVE_DIR / archived_name)

    uploaded_file.seek(0)
    live_path.write_bytes(uploaded_file.read())

    return True, "Database updated successfully."


def revert_to_default(db_key):
    """
    Overwrites the live file with the permanent default copy,
    discarding whatever was uploaded. The default copy itself is
    never modified, so this is always safe to repeat.

    Returns (ok: bool, message: str).
    """
    entry = DATABASE_REGISTRY[db_key]
    default_path = entry["default_path"]
    live_path = entry["live_path"]

    if not default_path.exists():
        return False, "No default copy found to revert to."

    shutil.copy2(default_path, live_path)
    return True, "Reverted to the default database."