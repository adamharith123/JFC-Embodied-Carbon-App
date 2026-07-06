from openpyxl import load_workbook


def validate_required_sheets(file_path, required_sheets):
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        available_sheets = wb.sheetnames
        missing = [sheet for sheet in required_sheets if sheet not in available_sheets]

        return {
            "valid": len(missing) == 0,
            "available_sheets": available_sheets,
            "missing_sheets": missing,
        }

    except Exception as e:
        return {
            "valid": False,
            "available_sheets": [],
            "missing_sheets": required_sheets,
            "error": str(e),
        }
