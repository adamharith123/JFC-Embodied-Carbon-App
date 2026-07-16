"""
Regenerate the app-only carbon database (ARUP_v2_app.xlsx) from the
master file (ARUP_v2.xlsx).

WHY THIS EXISTS
---------------
ARUP_v2.xlsx is the real, authored carbon database. Its "Apparatus
Output" sheet is formula-driven (SUMIF against a BOQ table), and its
used-range is heavily bloated with formatting far beyond the actual
data - reading it directly takes ~30-40s even though it only has ~40
real rows. The app doesn't need most of that file (it never reads
Estimate Calculator, Estimate Example, EPD Data, etc.) so this script
extracts just "Apparatus Output" and "dropdowns" - as their currently
CACHED calculated values, not live formulas - into a small, fast-
loading, static copy that the app actually reads.

HOW IT'S SAFE
-------------
This script only ever OPENS the master file in read-only mode. It
never re-saves ARUP_v2.xlsx. Re-saving it with openpyxl would wipe
the cached values of its formula-heavy sheets (openpyxl can't
recalculate formulas) - that's the mistake this script exists to
avoid repeating.

WHEN TO RUN THIS
----------------
Any time you've edited ARUP_v2.xlsx in Excel - added apparatus,
changed EC factors, updated the underlying BOQ data - and want the
app to see the changes. Just run:

    python regenerate_app_database.py

It'll print a before/after summary so you can see exactly what
changed before you commit the updated ARUP_v2_app.xlsx.
"""

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils.constants import CARBON_DATABASE_FILE, MASTER_CARBON_DATABASE_FILE

APPARATUS_SHEET_NAME = "Apparatus Output"
DROPDOWNS_SHEET_NAME = "dropdowns"

# Column index (0-based) of "Apparatus" and "Total (A1-3 + A4 + A5)"
# in the Apparatus Output sheet, used only for the diff summary below.
APPARATUS_NAME_COL = 2
APPARATUS_TOTAL_COL = 7


def extract_sheet_rows(worksheet, max_scan_rows):
    """
    Reads a worksheet's rows (as cached values, not formulas) up to
    max_scan_rows, then trims trailing rows that are entirely blank -
    i.e. it ignores the sheet's bloated declared dimensions and only
    keeps rows that actually have real content in them.
    """

    rows = list(
        worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True)
    )
    last_real_row_index = max(
        (i for i, row in enumerate(rows) if any(v is not None for v in row)),
        default=-1,
    )
    return rows[: last_real_row_index + 1]


def read_existing_apparatus_rows(app_database_path):
    """
    Reads the CURRENT (pre-regeneration) Apparatus Output rows from
    ARUP_v2_app.xlsx, if it already exists, so we can diff against
    what's about to be written. Returns {} if the file doesn't exist
    yet (e.g. first time this script is run).
    """

    if not app_database_path.exists():
        return {}

    wb = openpyxl.load_workbook(app_database_path, read_only=True, data_only=True)
    ws = wb[APPARATUS_SHEET_NAME]
    rows = extract_sheet_rows(ws, max_scan_rows=ws.max_row)
    wb.close()

    # rows[0] is the blank spacer row, rows[1] is the header row
    data_rows = rows[2:]
    return {
        row[APPARATUS_NAME_COL]: row[APPARATUS_TOTAL_COL]
        for row in data_rows
        if row[APPARATUS_NAME_COL] is not None
    }


def print_diff_summary(old_totals, new_data_rows):
    new_totals = {
        row[APPARATUS_NAME_COL]: row[APPARATUS_TOTAL_COL]
        for row in new_data_rows
        if row[APPARATUS_NAME_COL] is not None
    }

    def values_differ(a, b):
        # Ignore floating-point noise introduced by the read/write
        # round-trip itself (e.g. 170.38694 vs 170.38694000000004).
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            return round(a, 6) != round(b, 6)
        return a != b

    added = sorted(set(new_totals) - set(old_totals))
    removed = sorted(set(old_totals) - set(new_totals))
    changed = sorted(
        name
        for name in (set(new_totals) & set(old_totals))
        if values_differ(new_totals[name], old_totals[name])
    )

    print(f"\nApparatus Output: {len(old_totals)} -> {len(new_totals)} rows")

    if added:
        print(f"  Added ({len(added)}):")
        for name in added:
            print(f"    + {name}  (Total = {new_totals[name]})")

    if removed:
        print(f"  Removed ({len(removed)}):")
        for name in removed:
            print(f"    - {name}")

    if changed:
        print(f"  Changed Total value ({len(changed)}):")
        for name in changed:
            print(f"    ~ {name}: {old_totals[name]} -> {new_totals[name]}")

    if not (added or removed or changed):
        print("  No changes to apparatus data.")


def main():
    if not MASTER_CARBON_DATABASE_FILE.exists():
        print(f"ERROR: master file not found at {MASTER_CARBON_DATABASE_FILE}")
        sys.exit(1)

    print(f"Reading master file: {MASTER_CARBON_DATABASE_FILE.name}")
    old_totals = read_existing_apparatus_rows(CARBON_DATABASE_FILE)

    wb_read = openpyxl.load_workbook(
        MASTER_CARBON_DATABASE_FILE, read_only=True, data_only=True
    )

    ws_apparatus = wb_read[APPARATUS_SHEET_NAME]
    apparatus_rows = extract_sheet_rows(ws_apparatus, max_scan_rows=200)

    ws_dropdowns = wb_read[DROPDOWNS_SHEET_NAME]
    dropdown_rows = extract_sheet_rows(ws_dropdowns, max_scan_rows=200)

    wb_read.close()

    wb_new = openpyxl.Workbook()
    ws1 = wb_new.active
    ws1.title = APPARATUS_SHEET_NAME
    for row in apparatus_rows:
        ws1.append(row)

    ws2 = wb_new.create_sheet(DROPDOWNS_SHEET_NAME)
    for row in dropdown_rows:
        ws2.append(row)

    wb_new.save(CARBON_DATABASE_FILE)

    print(f"Saved: {CARBON_DATABASE_FILE.name}")
    print_diff_summary(old_totals, apparatus_rows[2:])


if __name__ == "__main__":
    main()
